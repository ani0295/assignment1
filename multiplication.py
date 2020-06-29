import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
import pandas as pd
import csv

wb = openpyxl.Workbook()
sheet = wb['Sheet'] # or wb.active
# sheet.column_dimensions['*'].width = 20
n = int(input('enter n  '))

boldFont = Font(bold=True)

for i in range(1, n + 1):
	sheet.cell(row=i + 1, column=1).value = i
	sheet.cell(row=i + 1, column=1).font =Font(size=24, bold=True, italic=False)
	sheet.cell(row=i + 1, column=1).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')
	sheet.cell(row=1, column=i+1).value = i
	sheet.cell(row=1, column=i+1).font = Font(size=24, bold=True, italic=False)
	sheet.cell(row=1, column=i+1).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')



for i in range(2, n + 2):
	for j in range(2, n + 2, 3):
	    sheet.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='8cff40', end_color='8cff40')
	    sheet.cell(row=i, column=j).font = Font(size=20, bold=True, italic=False)
for i in range(2, n + 2):
	for j in range(3, n + 2, 3):
	    sheet.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='ff6fff', end_color='ff6fff')
	    sheet.cell(row=i, column=j).font = Font(size=20, bold=True, italic=False)
for i in range(2, n + 2):
	for j in range(4, n + 2, 3):
	    sheet.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='76d7ea', end_color='76d7ea')
	    sheet.cell(row=i, column=j).font = Font(size=20, bold=True, italic=False)

for i in range(2, n + 2):
	for j in range(2, n + 2):
		x = sheet.cell(row=i, column=1).value
		y = sheet.cell(row=1, column=j).value
		sheet.cell(row=i, column=j).value = x * y


wb.save('multTable2.xlsx')
read_file = pd.read_excel (r'C:\\Users\\user\\Desktop\\python excel\\multTable2.xlsx')
read_file.to_csv (r'C:\\Users\\user\\Desktop\\python excel\\multTable2.csv', index = None, header=True)